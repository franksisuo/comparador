import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
from buscador import buscar_precios
import webbrowser
from openpyxl import Workbook

class ComparadorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Comparador de Precios")
        self.root.geometry("560x420")
        self.root.minsize(560, 420)

        self.productos = [
            "Perfil Cuadrado Acero 75x75x2 mm 6 m",
            "Perfil Cuadrado Acero 75x75x3 mm 6 m",
            "Perfil Cuadrado Acero 100x100x2 mm 6 m",
        ]
        self.ultimo_resultado = None

        # Etiqueta y selector para el producto
        ttk.Label(root, text="Producto:").pack(pady=10)
        self.producto_combo = ttk.Combobox(
            root,
            values=self.productos,
            width=45,
            state="readonly",
        )
        self.producto_combo.pack(pady=5)
        self.producto_combo.current(0)

        # Botón de buscar
        self.buscar_btn = ttk.Button(root, text="Buscar", command=self.iniciar_busqueda)
        self.buscar_btn.pack(pady=10)

        # Resultados
        self.resultados_frame = ttk.Frame(root)
        self.resultados_frame.pack(pady=20)

        ttk.Label(self.resultados_frame, text="Sodimac:").grid(row=0, column=0, sticky="w")
        self.sodimac_frame = ttk.Frame(self.resultados_frame)
        self.sodimac_frame.grid(row=0, column=1, sticky="w")
        self.sodimac_precio_label = tk.Label(self.sodimac_frame, text="...", justify='left', anchor='w')
        self.sodimac_precio_label.pack(side='left')
        self.sodimac_link_btn = tk.Button(self.sodimac_frame, text="Link", command=lambda: None)
        self.sodimac_link_btn.pack(side='left', padx=(10,0))

        ttk.Label(self.resultados_frame, text="Easy:").grid(row=1, column=0, sticky="w")
        self.easy_frame = ttk.Frame(self.resultados_frame)
        self.easy_frame.grid(row=1, column=1, sticky="w")
        self.easy_precio_label = tk.Label(self.easy_frame, text="...", justify='left', anchor='w')
        self.easy_precio_label.pack(side='left')
        self.easy_link_btn = tk.Button(self.easy_frame, text="Link", command=lambda: None)
        self.easy_link_btn.pack(side='left', padx=(10,0))

        ttk.Label(self.resultados_frame, text="Barraca Castro:").grid(row=2, column=0, sticky="w")
        self.castro_frame = ttk.Frame(self.resultados_frame)
        self.castro_frame.grid(row=2, column=1, sticky="w")
        self.castro_precio_label = tk.Label(self.castro_frame, text="...", justify='left', anchor='w')
        self.castro_precio_label.pack(side='left')
        self.castro_link_btn = tk.Button(self.castro_frame, text="Link", command=lambda: None)
        self.castro_link_btn.pack(side='left', padx=(10,0))

        self.exel_btn = ttk.Button(root, text="Exel", command=self.exportar_excel)
        self.exel_btn.pack(pady=5)
        self.exel_btn.pack_forget()

        # Barra de progreso
        self.progress = ttk.Progressbar(root, mode='indeterminate')
        self.progress.pack(pady=10, fill='x')
        self.estado_easy_label = ttk.Label(root, text="", foreground="#b26a00")
        self.estado_easy_label.pack(pady=(0, 5))

    def iniciar_busqueda(self):
        producto = self.producto_combo.get().strip()
        if not producto:
            return
        self.buscar_btn.config(state='disabled')
        self.progress.start()
        self.sodimac_precio_label.config(text="Buscando...")
        self.sodimac_link_btn.config(text="", state='disabled')
        self.easy_precio_label.config(text="Buscando...")
        self.easy_link_btn.config(text="", state='disabled')
        self.castro_precio_label.config(text="Buscando...")
        self.castro_link_btn.config(text="", state='disabled')
        self.estado_easy_label.config(text="")
        if self.exel_btn.winfo_ismapped():
            self.exel_btn.pack_forget()

        # Ejecutar en hilo separado
        threading.Thread(target=self.buscar, args=(producto,)).start()

    def buscar(self, producto):
        sodimac, easy, castro = buscar_precios(
            producto,
            on_status=lambda msg: self.root.after(0, self.mostrar_estado_easy, msg),
        )
        sodimac_precio, sodimac_url = sodimac if sodimac[0] else (None, None)
        easy_precio, easy_url = easy if easy[0] else (None, None)
        castro_precio, castro_url = castro if castro[0] else (None, None)
        self.root.after(
            0,
            self.actualizar_resultados,
            sodimac_precio,
            sodimac_url,
            easy_precio,
            easy_url,
            castro_precio,
            castro_url,
        )

    def mostrar_estado_easy(self, texto):
        self.estado_easy_label.config(text=texto)

    def actualizar_resultados(self, sodimac_precio, sodimac_url, easy_precio, easy_url, castro_precio, castro_url):
        self.sodimac_precio_label.config(text=f"Precio: {sodimac_precio or 'No encontrado'}")
        if sodimac_url:
            self.sodimac_link_btn.config(text="Link", state='normal', command=lambda: webbrowser.open(sodimac_url))
        else:
            self.sodimac_link_btn.config(text="", state='disabled')

        self.easy_precio_label.config(text=f"Precio: {easy_precio or 'No encontrado'}")
        if easy_url:
            self.easy_link_btn.config(text="Link", state='normal', command=lambda: webbrowser.open(easy_url))
        else:
            self.easy_link_btn.config(text="", state='disabled')

        self.castro_precio_label.config(text=f"Precio: {castro_precio or 'No encontrado'}")
        if castro_url:
            self.castro_link_btn.config(text="Link", state='normal', command=lambda: webbrowser.open(castro_url))
        else:
            self.castro_link_btn.config(text="", state='disabled')

        self.ultimo_resultado = {
            "producto": self.producto_combo.get().strip(),
            "easy": easy_precio or "No encontrado",
            "sodimac": sodimac_precio or "No encontrado",
            "castro": castro_precio or "No encontrado",
        }

        self.buscar_btn.config(state='normal')
        self.progress.stop()
        self.estado_easy_label.config(text="")
        if not self.exel_btn.winfo_ismapped():
            self.exel_btn.pack(pady=5)

    def exportar_excel(self):
        if not self.ultimo_resultado:
            messagebox.showwarning("Sin datos", "Primero debes realizar una busqueda.")
            return

        ruta = filedialog.asksaveasfilename(
            title="Guardar Excel",
            defaultextension=".xlsx",
            filetypes=[("Archivo Excel", "*.xlsx")],
            initialfile="comparador_precios.xlsx",
        )
        if not ruta:
            return

        wb = Workbook()
        ws = wb.active
        ws.title = "Comparador"
        ws["A1"] = "Producto"
        ws["B1"] = "Easy"
        ws["C1"] = "Sodimac"
        ws["D1"] = "Barraca Castro"
        ws["A2"] = self.ultimo_resultado["producto"]
        ws["B2"] = self.ultimo_resultado["easy"]
        ws["C2"] = self.ultimo_resultado["sodimac"]
        ws["D2"] = self.ultimo_resultado["castro"]
        wb.save(ruta)
        messagebox.showinfo("Listo", f"Excel guardado en:\n{ruta}")

if __name__ == "__main__":
    root = tk.Tk()
    app = ComparadorApp(root)
    root.mainloop()
